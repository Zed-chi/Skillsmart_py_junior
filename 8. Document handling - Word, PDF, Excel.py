"""
Document handling task
 Word: docx module
 Excel: openpyxl module
 PDF: PyPDF2 module

"""


def replace_string_in_doc():
    from docx import Document
    import re
    file = input("Открыть файл:")+".docx"
    string_to_replace = input("Требуется Заменить:")
    pat =re.compile(r"(\W|\s)"+string_to_replace+r"(\W|\s)")
    replace = input("На:")
    d = Document(file)
    for i in d.paragraphs:
        if pat.search(i.text) is not None:
            i.text = re.sub(pat, replace, i.text)
    d.save(file)

def fill_table_with_random():
    from openpyxl import Workbook, load_workbook
    import random
    file = Workbook()
    name = input("page_num:")+".xlsx"
    sheet = file.create_sheet(title="Тестовый лист")
    for i in range(1,13):
        for j in range(1,13):
            sheet.cell(column = j, row = i, value = random.randint(0, 12345))
    file.save(filename = name)

def get_text_from_pdf():
    import PyPDF2
    name = input("Открыть файл:")+".pdf"
    file = open(name, 'rb')
    pdf = PyPDF2.PdfFileReader(file)
    # извлекаем и раскодировываем её содержимое:
    while(True):
        page_num = input("№ стр(0..)/Выход(Q):")
        if page_num == "Q":
            print("Конец")
            break
        try:
            page = pdf.getPage(int(page_num))
            page_content = page.extractText()
            print(page_content.encode('utf-8'))
        except:
            print("Некорректный ввод")
            continue

replace_string_in_doc()
##fill_table_with_random()
##get_text_from_pdf()
